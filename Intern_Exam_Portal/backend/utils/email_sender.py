"""
Email utility for sending candidate test links using Gmail SMTP.
Uses SMTP_EMAIL and SMTP_APP_PASSWORD from environment/.env file.
"""
import smtplib
import os
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from dotenv import load_dotenv

# Load .env from the backend root directory regardless of where we're called from
_env_path = Path(__file__).resolve().parent.parent / ".env"
load_dotenv(dotenv_path=_env_path, override=True)

SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_EMAIL = os.getenv("SMTP_EMAIL", "")
SMTP_APP_PASSWORD = os.getenv("SMTP_APP_PASSWORD", "")
SENDER_NAME = os.getenv("SENDER_NAME", "InternAssess Portal")


def send_test_link_email(
    candidate_name: str,
    candidate_email: str,
    test_link: str,
    expires_at: str,
    job_position: str = "",
    assessment_title: str = "",
    experience_level: str = "",
) -> bool:
    """
    Send a test link email to the candidate.
    Returns True on success, False on failure.
    """
    if not SMTP_EMAIL or not SMTP_APP_PASSWORD:
        print("[WARN] SMTP not configured. Email not sent. Set SMTP_EMAIL and SMTP_APP_PASSWORD in .env")
        return False

    exp_label = "Experienced" if experience_level == "experienced" else "Fresher"
    role_line = f"{job_position} ({exp_label})" if job_position else "Assessment"
    subject_title = assessment_title if assessment_title else "InternAssess MCQ Test"

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Your {subject_title} Test Link — {role_line}"
        msg["From"] = f"{SENDER_NAME} <{SMTP_EMAIL}>"
        msg["To"] = candidate_email

        # Plain text fallback
        plain = f"""
Hello {candidate_name},

You have been invited to take the {subject_title} for the role of {role_line}.

Your unique test link (valid for 1 hour):
{test_link}

Expires at: {expires_at}

Instructions:
- Open the link on a Desktop or Laptop browser only
- Do NOT switch tabs or minimize the window during the test
- Answer all questions before submitting
- You cannot re-take the test once submitted

Good luck!
InternAssess Team
        """.strip()

        # HTML version
        html = f"""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f5f7fb;font-family:Inter,Arial,sans-serif;">
  <div style="max-width:600px;margin:40px auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);">
    <!-- Header -->
    <div style="background:linear-gradient(135deg,#1a56db,#1e40af);padding:36px 40px;text-align:center;">
      <div style="display:inline-block;background:rgba(255,255,255,0.15);border-radius:10px;padding:10px 18px;margin-bottom:16px;">
        <span style="color:#fff;font-size:18px;font-weight:800;letter-spacing:1px;">IA</span>
      </div>
      <h1 style="color:#fff;margin:0;font-size:26px;font-weight:800;">InternAssess</h1>
      <p style="color:rgba(255,255,255,0.8);margin:8px 0 0;font-size:14px;">Internship Assessment Portal</p>
    </div>

    <!-- Body -->
    <div style="padding:40px;">
      <h2 style="font-size:20px;color:#111827;margin-bottom:20px;">Hello, {candidate_name}!</h2>

      <p style="color:#6b7280;font-size:15px;line-height:1.6;margin-bottom:28px;">
        You have been invited to take the <strong>{subject_title}</strong> for the role of
        <strong>{role_line}</strong>.
        Click the button below to begin your assessment.
      </p>

      <!-- CTA Button -->
      <div style="text-align:center;margin-bottom:28px;">
        <a href="{test_link}"
           style="display:inline-block;background:#1a56db;color:#fff;padding:14px 36px;border-radius:8px;font-size:16px;font-weight:700;text-decoration:none;letter-spacing:0.3px;">
          Start My Test →
        </a>
      </div>

      <!-- Link box -->
      <div style="background:#f5f7fb;border:1px solid #e5e7eb;border-radius:8px;padding:14px 16px;margin-bottom:20px;">
        <p style="font-size:12px;color:#6b7280;margin:0 0 6px;text-transform:uppercase;letter-spacing:0.05em;font-weight:600;">Or copy this link:</p>
        <p style="font-size:13px;color:#1a56db;margin:0;word-break:break-all;">{test_link}</p>
      </div>

      <!-- Expiry -->
      <div style="background:#fff7ed;border:1px solid #fed7aa;border-radius:8px;padding:12px 16px;margin-bottom:24px;">
        <p style="font-size:13px;color:#9a3412;margin:0;">
          <strong>Expires at:</strong> {expires_at} IST &nbsp;|&nbsp; <strong>Valid for 1 hour only</strong>
        </p>
      </div>

      <!-- Rules -->
      <h3 style="font-size:14px;color:#374151;margin-bottom:10px;">Important Instructions:</h3>
      <ul style="color:#6b7280;font-size:14px;line-height:1.8;padding-left:20px;">
        <li>Open the link on a <strong>Desktop or Laptop</strong> only (mobiles/tablets are blocked)</li>
        <li>Do <strong>NOT</strong> switch tabs or minimize the window during the test</li>
        <li>Answer all questions before submitting</li>
        <li>The test <strong>auto-submits</strong> on the 3rd tab switch</li>
        <li>You cannot re-take the test once submitted</li>
      </ul>
    </div>

    <!-- Footer -->
    <div style="background:#f9fafb;border-top:1px solid #e5e7eb;padding:20px 40px;text-align:center;">
      <p style="font-size:12px;color:#9ca3af;margin:0;">
        This is an automated message from InternAssess. Please do not reply to this email.
      </p>
    </div>
  </div>
</body>
</html>
        """.strip()

        msg.attach(MIMEText(plain, "plain"))
        msg.attach(MIMEText(html, "html"))

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.ehlo()
            server.starttls()
            server.login(SMTP_EMAIL, SMTP_APP_PASSWORD)
            server.sendmail(SMTP_EMAIL, candidate_email, msg.as_string())

        print(f"[OK] Email sent to {candidate_email} ({role_line})")
        return True

    except Exception as e:
        print(f"[ERROR] Failed to send email to {candidate_email}: {e}")
        return False


def send_assessment_report_email(
    admin_email: str,
    assessment_title: str,
    job_position: str,
    db,
    assessment_id: int,
) -> bool:
    """
    Send a completion notification email to the admin who created the assessment,
    with an Excel batch report of all candidates as an attachment.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from email.mime.base import MIMEBase
    from email import encoders
    from datetime import timezone, timedelta

    if not SMTP_EMAIL or not SMTP_APP_PASSWORD:
        print("[WARN] SMTP not configured. Report email not sent.")
        return False

    IST = timezone(timedelta(hours=5, minutes=30))

    try:
        # ── Build Excel report ────────────────────────────────────────────────
        from utils.grading import grade_and_rank_candidates
        all_candidates = grade_and_rank_candidates(db)
        batch = [c for c in all_candidates if c.get("assessment_title") == assessment_title]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Batch Report"

        # Styles
        header_font    = Font(bold=True, color="FFFFFF", size=11)
        header_fill    = PatternFill("solid", fgColor="1A56DB")
        center_align   = Alignment(horizontal="center", vertical="center")
        thin_border    = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        submitted_fill = PatternFill("solid", fgColor="ECFDF5")
        pending_fill   = PatternFill("solid", fgColor="FFFBEB")
        absent_fill    = PatternFill("solid", fgColor="FEF2F2")

        headers = ["#", "Name", "Email", "Score", "Percentage (%)", "Rank",
                   "Tab Switches", "Status", "Submitted At"]
        col_widths = [5, 25, 32, 10, 16, 8, 14, 14, 22]

        for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = w

        ws.row_dimensions[1].height = 22

        # Sort: submitted first by rank, then pending, then not attended
        def sort_key(c):
            if c["status"] == "submitted":   return (0, c.get("rank") or 9999)
            if c["status"] == "pending":     return (1, 9999)
            return (2, 9999)

        batch_sorted = sorted(batch, key=sort_key)

        for row_idx, c in enumerate(batch_sorted, 2):
            submitted_at_str = ""
            if c.get("submitted_at"):
                try:
                    dt = c["submitted_at"]
                    if hasattr(dt, "replace"):
                        dt_ist = dt.replace(tzinfo=timezone.utc).astimezone(IST)
                        submitted_at_str = dt_ist.strftime("%Y-%m-%d %I:%M %p IST")
                except Exception:
                    submitted_at_str = str(c.get("submitted_at", ""))

            score_str = f"{c['score']}/{c['total']}" if c.get("score") is not None else "—"
            pct_str   = f"{c['percentage']:.1f}" if c.get("percentage") is not None else "—"
            rank_str  = str(c["rank"]) if c.get("rank") else "—"
            status    = c.get("status", "—").replace("_", " ").title()

            row_data = [
                row_idx - 1,
                c["name"],
                c["email"],
                score_str,
                pct_str,
                rank_str,
                c.get("tab_switches", 0),
                status,
                submitted_at_str,
            ]

            if c["status"] == "submitted":    row_fill = submitted_fill
            elif c["status"] == "pending":    row_fill = pending_fill
            else:                             row_fill = absent_fill

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="center" if col_idx in (1, 4, 5, 6, 7, 8) else "left",
                                           vertical="center")

        # Summary row
        summary_row = len(batch_sorted) + 2
        total_submitted = sum(1 for c in batch if c["status"] == "submitted")
        total_pending   = sum(1 for c in batch if c["status"] == "pending")
        total_absent    = sum(1 for c in batch if c["status"] == "not_attended")
        ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
        ws.cell(row=summary_row, column=2,
                value=f"Total: {len(batch)} | Submitted: {total_submitted} | Pending: {total_pending} | Not Attended: {total_absent}")

        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        safe_title = "".join(c if c.isalnum() or c in " _-" else "_" for c in assessment_title)
        filename = f"{safe_title}_Batch_Report.xlsx"

        # ── Build email ───────────────────────────────────────────────────────
        msg = MIMEMultipart("mixed")
        msg["Subject"] = f"[InternAssess] New Submission — {assessment_title} ({job_position})"
        msg["From"]    = f"{SENDER_NAME} <{SMTP_EMAIL}>"
        msg["To"]      = admin_email

        plain_body = f"""
Hello Admin,

A candidate has completed the assessment for:

  Assessment : {assessment_title}
  Role       : {job_position}

Please find the full batch report attached to this email.

— InternAssess Portal
        """.strip()

        html_body = f"""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f5f7fb;font-family:Inter,Arial,sans-serif;">
  <div style="max-width:600px;margin:40px auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);">
    <div style="background:linear-gradient(135deg,#1a56db,#1e40af);padding:36px 40px;text-align:center;">
      <div style="display:inline-block;background:rgba(255,255,255,0.15);border-radius:10px;padding:10px 18px;margin-bottom:16px;">
        <span style="color:#fff;font-size:18px;font-weight:800;letter-spacing:1px;">IA</span>
      </div>
      <h1 style="color:#fff;margin:0;font-size:24px;font-weight:800;">InternAssess</h1>
      <p style="color:rgba(255,255,255,0.8);margin:8px 0 0;font-size:13px;">Admin Notification</p>
    </div>
    <div style="padding:40px;">
      <h2 style="font-size:18px;color:#111827;margin-bottom:8px;">New Test Submission Received</h2>
      <p style="color:#6b7280;font-size:14px;line-height:1.6;margin-bottom:24px;">
        A candidate has completed the assessment. The full batch report is attached to this email.
      </p>
      <div style="background:#f5f7fb;border:1px solid #e5e7eb;border-radius:8px;padding:18px 20px;margin-bottom:24px;">
        <table style="width:100%;font-size:14px;border-collapse:collapse;">
          <tr>
            <td style="color:#6b7280;padding:5px 0;width:40%;">Assessment</td>
            <td style="color:#111827;font-weight:600;">{assessment_title}</td>
          </tr>
          <tr>
            <td style="color:#6b7280;padding:5px 0;">Role</td>
            <td style="color:#111827;font-weight:600;">{job_position}</td>
          </tr>
          <tr>
            <td style="color:#6b7280;padding:5px 0;">Total Candidates</td>
            <td style="color:#111827;font-weight:600;">{len(batch)}</td>
          </tr>
          <tr>
            <td style="color:#6b7280;padding:5px 0;">Submitted</td>
            <td style="color:#059669;font-weight:600;">{total_submitted}</td>
          </tr>
          <tr>
            <td style="color:#6b7280;padding:5px 0;">Pending</td>
            <td style="color:#d97706;font-weight:600;">{total_pending}</td>
          </tr>
          <tr>
            <td style="color:#6b7280;padding:5px 0;">Not Attended</td>
            <td style="color:#dc2626;font-weight:600;">{total_absent}</td>
          </tr>
        </table>
      </div>
      <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:14px 16px;">
        <p style="font-size:13px;color:#1e40af;margin:0;">
          📎 The complete batch report (Excel) is attached to this email with scores, rankings, and tab-switch details for all candidates.
        </p>
      </div>
    </div>
    <div style="background:#f9fafb;border-top:1px solid #e5e7eb;padding:20px 40px;text-align:center;">
      <p style="font-size:12px;color:#9ca3af;margin:0;">
        This is an automated notification from InternAssess. Do not reply to this email.
      </p>
    </div>
  </div>
</body>
</html>
        """.strip()

        alt_part = MIMEMultipart("alternative")
        alt_part.attach(MIMEText(plain_body, "plain"))
        alt_part.attach(MIMEText(html_body, "html"))
        msg.attach(alt_part)

        # Attach Excel file
        attachment = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        attachment.set_payload(excel_buffer.read())
        encoders.encode_base64(attachment)
        attachment.add_header("Content-Disposition", "attachment", filename=filename)
        msg.attach(attachment)

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.ehlo()
            server.starttls()
            server.login(SMTP_EMAIL, SMTP_APP_PASSWORD)
            server.sendmail(SMTP_EMAIL, admin_email, msg.as_string())

        print(f"[OK] Assessment report sent to admin: {admin_email}")
        return True

    except Exception as e:
        print(f"[ERROR] Failed to send report email to {admin_email}: {e}")
        return False


def send_reminder_email(
    candidate_name: str,
    candidate_email: str,
    test_link: str,
    expires_at: str,
    job_position: str = "",
    assessment_title: str = "",
) -> bool:
    """
    Send a reminder email to a candidate whose test link is expiring soon
    and who has not yet started the test.
    """
    if not SMTP_EMAIL or not SMTP_APP_PASSWORD:
        print("[WARN] SMTP not configured. Reminder email not sent.")
        return False

    subject_title = assessment_title if assessment_title else "InternAssess MCQ Test"
    role_line = job_position if job_position else "the assessment"

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Reminder: Your test link for {subject_title} is expiring soon!"
        msg["From"]    = f"{SENDER_NAME} <{SMTP_EMAIL}>"
        msg["To"]      = candidate_email

        plain = f"""
Hello {candidate_name},

This is a reminder that your test link for {subject_title} ({role_line}) is expiring soon.

Your test link:
{test_link}

Expires at: {expires_at} IST

Please attempt the test before the link expires.

Instructions:
- Open the link on a Desktop or Laptop browser only
- Do NOT switch tabs or minimize the window during the test
- Answer all questions before submitting
- You cannot re-take the test once submitted

Good luck!
InternAssess Team
        """.strip()

        html = f"""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f5f7fb;font-family:Inter,Arial,sans-serif;">
  <div style="max-width:600px;margin:40px auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);">

    <!-- Header -->
    <div style="background:linear-gradient(135deg,#d97706,#b45309);padding:36px 40px;text-align:center;">
      <div style="display:inline-block;background:rgba(255,255,255,0.15);border-radius:10px;padding:10px 18px;margin-bottom:16px;">
        <span style="color:#fff;font-size:18px;font-weight:800;letter-spacing:1px;">IA</span>
      </div>
      <h1 style="color:#fff;margin:0;font-size:24px;font-weight:800;">InternAssess</h1>
      <p style="color:rgba(255,255,255,0.85);margin:8px 0 0;font-size:13px;">Test Link Expiry Reminder</p>
    </div>

    <!-- Body -->
    <div style="padding:40px;">
      <div style="background:#fffbeb;border:1px solid #fcd34d;border-radius:8px;padding:14px 18px;margin-bottom:24px;display:flex;align-items:center;gap:10px;">
        <span style="font-size:20px;">⏰</span>
        <p style="font-size:14px;color:#92400e;margin:0;font-weight:600;">
          Your test link is expiring soon — please attempt it before it expires!
        </p>
      </div>

      <h2 style="font-size:18px;color:#111827;margin-bottom:8px;">Hello, {candidate_name}!</h2>
      <p style="color:#6b7280;font-size:14px;line-height:1.6;margin-bottom:24px;">
        You have a pending assessment — <strong>{subject_title}</strong> for the role of
        <strong>{role_line}</strong>. You haven't attempted it yet and your link is expiring soon.
      </p>

      <!-- Expiry warning -->
      <div style="background:#fef2f2;border:1px solid #fecaca;border-radius:8px;padding:12px 16px;margin-bottom:24px;">
        <p style="font-size:13px;color:#991b1b;margin:0;">
          <strong>⚠️ Expires at:</strong> {expires_at} IST
        </p>
      </div>

      <!-- CTA Button -->
      <div style="text-align:center;margin-bottom:28px;">
        <a href="{test_link}"
           style="display:inline-block;background:#d97706;color:#fff;padding:14px 36px;border-radius:8px;font-size:16px;font-weight:700;text-decoration:none;letter-spacing:0.3px;">
          Attempt My Test Now →
        </a>
      </div>

      <!-- Link box -->
      <div style="background:#f5f7fb;border:1px solid #e5e7eb;border-radius:8px;padding:14px 16px;margin-bottom:24px;">
        <p style="font-size:12px;color:#6b7280;margin:0 0 6px;text-transform:uppercase;letter-spacing:0.05em;font-weight:600;">Or copy this link:</p>
        <p style="font-size:13px;color:#d97706;margin:0;word-break:break-all;">{test_link}</p>
      </div>

      <!-- Rules -->
      <h3 style="font-size:14px;color:#374151;margin-bottom:10px;">Important Instructions:</h3>
      <ul style="color:#6b7280;font-size:14px;line-height:1.8;padding-left:20px;">
        <li>Open the link on a <strong>Desktop or Laptop</strong> only</li>
        <li>Do <strong>NOT</strong> switch tabs or minimize the window during the test</li>
        <li>The test <strong>auto-submits</strong> on the 3rd tab switch</li>
        <li>You cannot re-take the test once submitted</li>
      </ul>
    </div>

    <!-- Footer -->
    <div style="background:#f9fafb;border-top:1px solid #e5e7eb;padding:20px 40px;text-align:center;">
      <p style="font-size:12px;color:#9ca3af;margin:0;">
        This is an automated reminder from InternAssess. Please do not reply to this email.
      </p>
    </div>
  </div>
</body>
</html>
        """.strip()

        msg.attach(MIMEText(plain, "plain"))
        msg.attach(MIMEText(html, "html"))

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.ehlo()
            server.starttls()
            server.login(SMTP_EMAIL, SMTP_APP_PASSWORD)
            server.sendmail(SMTP_EMAIL, candidate_email, msg.as_string())

        print(f"[OK] Reminder email sent to {candidate_email}")
        return True

    except Exception as e:
        print(f"[ERROR] Failed to send reminder email to {candidate_email}: {e}")
        return False


def send_candidate_result_email(
    candidate_name: str,
    candidate_email: str,
    job_position: str,
    assessment_title: str,
    decision: str,          # 'selected' | 'rejected'
    score=None,
    total=None,
    percentage=None,
) -> bool:
    """
    Send a Selected or Rejected result email to the candidate.
    Triggered manually by admin from the CandidateDetail page.
    """
    if not SMTP_EMAIL or not SMTP_APP_PASSWORD:
        print("[WARN] SMTP not configured. Result email not sent.")
        return False

    is_selected = decision == "selected"
    subject_title = assessment_title if assessment_title else "Assessment"
    role_line = job_position if job_position else "the position"

    if is_selected:
        subject = f"Congratulations! You have been Selected — {subject_title}"
        header_bg = "linear-gradient(135deg,#059669,#047857)"
        header_label = "Selected ✓"
        headline = f"Congratulations, {candidate_name}!"
        body_para = (
            f"We are pleased to inform you that you have been <strong>selected</strong> "
            f"for the role of <strong>{role_line}</strong> based on your performance in "
            f"the <strong>{subject_title}</strong>."
        )
        next_steps = (
            "Our team will reach out to you shortly with the next steps. "
            "Please keep an eye on your inbox and be ready to respond promptly."
        )
        banner_bg = "#ecfdf5"
        banner_border = "#6ee7b7"
        banner_color = "#065f46"
        banner_icon = "🎉"
        banner_text = "You have been selected! Well done."
    else:
        subject = f"Update on your Application — {subject_title}"
        header_bg = "linear-gradient(135deg,#dc2626,#b91c1c)"
        header_label = "Application Update"
        headline = f"Hello, {candidate_name}"
        body_para = (
            f"Thank you for taking the time to appear for the <strong>{subject_title}</strong> "
            f"for the role of <strong>{role_line}</strong>. After careful evaluation, "
            f"we regret to inform you that we are unable to move forward with your application at this time."
        )
        next_steps = (
            "We appreciate your effort and encourage you to keep building your skills. "
            "We wish you all the best in your future endeavors."
        )
        banner_bg = "#fef2f2"
        banner_border = "#fca5a5"
        banner_color = "#991b1b"
        banner_icon = "📋"
        banner_text = "Thank you for participating in our assessment process."

    score_section = ""
    if score is not None and total is not None and percentage is not None:
        score_section = f"""
      <div style="background:#f5f7fb;border:1px solid #e5e7eb;border-radius:8px;padding:16px 20px;margin-bottom:24px;">
        <p style="font-size:13px;color:#6b7280;margin:0 0 10px;text-transform:uppercase;
                  letter-spacing:0.05em;font-weight:600;">Your Score</p>
        <table style="width:100%;font-size:14px;border-collapse:collapse;">
          <tr>
            <td style="color:#6b7280;padding:4px 0;">Score</td>
            <td style="color:#111827;font-weight:600;text-align:right;">{score} / {total}</td>
          </tr>
          <tr>
            <td style="color:#6b7280;padding:4px 0;">Percentage</td>
            <td style="color:#111827;font-weight:600;text-align:right;">{percentage:.1f}%</td>
          </tr>
        </table>
      </div>"""

    plain = f"""
Hello {candidate_name},

{'Congratulations!' if is_selected else 'Thank you for your participation.'}

{body_para.replace('<strong>', '').replace('</strong>', '')}

{f'Your Score: {score}/{total} ({percentage:.1f}%)' if score is not None else ''}

{next_steps}

Best regards,
InternAssess Team
    """.strip()

    html = f"""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f5f7fb;font-family:Inter,Arial,sans-serif;">
  <div style="max-width:600px;margin:40px auto;background:#fff;border-radius:12px;
              overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);">
    <div style="background:{header_bg};padding:36px 40px;text-align:center;">
      <div style="display:inline-block;background:rgba(255,255,255,0.15);border-radius:10px;
                  padding:10px 18px;margin-bottom:16px;">
        <span style="color:#fff;font-size:18px;font-weight:800;letter-spacing:1px;">IA</span>
      </div>
      <h1 style="color:#fff;margin:0;font-size:26px;font-weight:800;">InternAssess</h1>
      <p style="color:rgba(255,255,255,0.85);margin:10px 0 0;font-size:15px;font-weight:600;">
        {header_label}
      </p>
    </div>
    <div style="padding:40px;">
      <div style="background:{banner_bg};border:1px solid {banner_border};border-radius:8px;
                  padding:14px 18px;margin-bottom:24px;">
        <p style="font-size:14px;color:{banner_color};margin:0;font-weight:600;">
          {banner_icon} {banner_text}
        </p>
      </div>
      <h2 style="font-size:20px;color:#111827;margin-bottom:16px;">{headline}</h2>
      <p style="color:#6b7280;font-size:15px;line-height:1.7;margin-bottom:24px;">
        {body_para}
      </p>
      {score_section}
      <p style="color:#6b7280;font-size:14px;line-height:1.7;margin-bottom:24px;">
        {next_steps}
      </p>
      <p style="color:#6b7280;font-size:14px;line-height:1.7;">
        Best regards,<br><strong style="color:#111827;">InternAssess Team</strong>
      </p>
    </div>
    <div style="background:#f9fafb;border-top:1px solid #e5e7eb;padding:20px 40px;text-align:center;">
      <p style="font-size:12px;color:#9ca3af;margin:0;">
        This is an official communication from InternAssess. Please do not reply to this email.
      </p>
    </div>
  </div>
</body>
</html>
    """.strip()

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = f"{SENDER_NAME} <{SMTP_EMAIL}>"
        msg["To"] = candidate_email
        msg.attach(MIMEText(plain, "plain"))
        msg.attach(MIMEText(html, "html"))

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.ehlo()
            server.starttls()
            server.login(SMTP_EMAIL, SMTP_APP_PASSWORD)
            server.sendmail(SMTP_EMAIL, candidate_email, msg.as_string())

        print(f"[OK] Result email ({decision}) sent to {candidate_email}")
        return True
    except Exception as e:
        print(f"[ERROR] Failed to send result email to {candidate_email}: {e}")
        return False
