"""
Generate a sample MCQ Excel file for testing the portal.
Run: python generate_sample_excel.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "MCQs"

# Header
headers = ['subject', 'question', 'option_a', 'option_b', 'option_c', 'option_d', 'correct_answer', 'question_mark']
header_fill = PatternFill("solid", fgColor="1A56DB")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Sample 25 MCQs — columns: (subject, question, option_a, option_b, option_c, option_d, correct_answer)
questions = [
    ("HTML",       "What does HTML stand for?", "HyperText Markup Language", "HighText Machine Language", "HyperText and links Markup Language", "None of the above", "a"),
    ("CSS",        "Which language is used for styling web pages?", "HTML", "CSS", "JavaScript", "Python", "b"),
    ("CSS",        "What does CSS stand for?", "Computer Style Sheets", "Creative Style Sheets", "Cascading Style Sheets", "Colorful Style Sheets", "c"),
    ("HTML",       "Which HTML tag is used for the largest heading?", "<h6>", "<heading>", "<h1>", "<head>", "c"),
    ("HTML",       "What is the correct HTML element for inserting a line break?", "<br>", "<lb>", "<break>", "<newline>", "a"),
    ("Python",     "Which symbol is used for comments in Python?", "//", "/*", "#", "--", "c"),
    ("SQL",        "What does SQL stand for?", "Structured Query Language", "Simple Query Language", "Standard Question Language", "Structured Question Language", "a"),
    ("Python",     "Which data type stores decimal numbers in Python?", "int", "float", "str", "bool", "b"),
    ("SQL",        "What is a primary key in a database?", "A key for locking the database", "A unique identifier for each record", "The first column", "An index column", "b"),
    ("Networking", "Which HTTP method is used to send data to a server?", "GET", "DELETE", "POST", "PUT", "c"),
    ("API",        "What does API stand for?", "Application Programming Integration", "Application Protocol Interface", "Application Programming Interface", "Automated Program Interface", "c"),
    ("JavaScript", "Which of the following is NOT a JavaScript framework?", "React", "Angular", "Vue", "Django", "d"),
    ("API",        "What is a REST API?", "A type of database", "A protocol for file transfer", "An architectural style for APIs over HTTP", "A JavaScript framework", "c"),
    ("Programming","What does OOP stand for?", "Object Oriented Programming", "Open Oriented Protocol", "Object Output Process", "None of the above", "a"),
    ("Python",     "Which Python keyword is used to create a function?", "function", "define", "def", "func", "c"),
    ("Tools",      "What is Git used for?", "Database management", "Version control", "Web development", "API testing", "b"),
    ("HTML",       "Which tag is used to create a hyperlink in HTML?", "<link>", "<a>", "<href>", "<url>", "b"),
    ("Algorithm",  "What is the time complexity of binary search?", "O(n)", "O(n²)", "O(log n)", "O(1)", "c"),
    ("SQL",        "Which of these is a NoSQL database?", "MySQL", "PostgreSQL", "MongoDB", "SQLite", "c"),
    ("API",        "What does JSON stand for?", "JavaScript Object Numbers", "JavaScript Object Notation", "Java Source Object Network", "JavaScript Output Notation", "b"),
    ("Python",     "Which command is used to install Python packages?", "npm install", "pip install", "apt install", "brew install", "b"),
    ("Python",     "What is a virtual environment in Python?", "A cloud server", "An isolated Python environment", "A Docker container", "A virtual machine", "b"),
    ("Networking", "Which HTTP status code means 'Not Found'?", "200", "401", "500", "404", "d"),
    ("SQL",        "What is the purpose of an index in a database?", "To encrypt data", "To backup data", "To speed up data retrieval", "To delete data", "c"),
    ("Python",     "Which of the following is a Python web framework?", "React", "Angular", "FastAPI", "Vue", "c"),
]

for row, (subject, q, a, b, c, d, ans) in enumerate(questions, 2):
    ws.cell(row=row, column=1, value=subject)
    ws.cell(row=row, column=2, value=q)
    ws.cell(row=row, column=3, value=a)
    ws.cell(row=row, column=4, value=b)
    ws.cell(row=row, column=5, value=c)
    ws.cell(row=row, column=6, value=d)
    ws.cell(row=row, column=7, value=ans)
    # question_mark stays blank for MCQ rows (column 8)

# Descriptive questions — columns: (subject, question, question_mark)
# Options + correct_answer stay blank for descriptive rows
descriptive_questions = [
    ("Python",      "Describe Object-Oriented Programming in Python and its four main principles.", "5M"),
    ("SQL",         "Explain the difference between INNER JOIN and LEFT JOIN with an example.",     "5M"),
    ("Programming", "What is the difference between a stack and a queue? Give a real-world example for each.", "2M"),
    ("Algorithm",   "Describe the steps involved in the Software Development Life Cycle (SDLC).",   "10M"),
]

start_row = len(questions) + 2  # after MCQ rows
for offset, (subject, q, mark) in enumerate(descriptive_questions):
    row = start_row + offset
    ws.cell(row=row, column=1, value=subject)
    ws.cell(row=row, column=2, value=q)
    ws.cell(row=row, column=8, value=mark)
    # columns 3-7 (options + correct_answer) stay blank

# Column widths
ws.column_dimensions['A'].width = 14   # subject
ws.column_dimensions['B'].width = 60   # question
for col in ['C','D','E','F']:
    ws.column_dimensions[col].width = 35
ws.column_dimensions['G'].width = 18   # correct_answer
ws.column_dimensions['H'].width = 14   # question_mark

filename = "sample_mcqs.xlsx"
wb.save(filename)
print(f"OK Sample Excel file created: {filename}")
print(f"   {len(questions)} MCQ rows + {len(descriptive_questions)} descriptive rows.")
print(f"   Upload this file on the Admin > Upload MCQs page.")
